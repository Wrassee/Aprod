#!/usr/bin/env python3
"""
Egyesített Excel template létrehozása - minden kérdéstípus egy helyen
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys

def create_unified_template():
    """Létrehoz egy egyesített Excel template-et minden kérdéstípussal"""
    
    # Új munkafüzet létrehozása
    wb = openpyxl.Workbook()
    
    # Táblázat átnevezése
    ws = wb.active
    ws.title = "questions"
    
    # Fejléc stílusa
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Fejléc sorának beállítása
    headers = [
        "id", "title_hu", "title_de", "type", "required", 
        "cell_reference", "group", "group_de", "order", 
        "unit", "min_value", "max_value", "calculation_formula", "calculation_inputs"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Minta adatok hozzáadása - minden kérdéstípus
    questions = [
        # Szöveges kérdések
        ["1", "Átvevő neve", "Name des Empfängers", "text", "TRUE", "F9", 
         "Általános adatok", "Allgemeine Daten", "1", "", "", "", "", ""],
        ["2", "Létesítmény címe", "Anschrift der Anlage", "text", "TRUE", "Q9", 
         "Általános adatok", "Allgemeine Daten", "2", "", "", "", "", ""],
        ["3", "Emelet szám", "Anzahl Stockwerke", "number", "TRUE", "G13", 
         "Általános adatok", "Allgemeine Daten", "3", "", "", "", "", ""],
        ["4", "Modernizációs partner", "Modernisierungspartner", "text", "TRUE", "O13", 
         "Általános adatok", "Allgemeine Daten", "4", "", "", "", "", ""],
        ["5", "OTIS szerződésszám", "OTIS Vertragsnummer", "text", "TRUE", "G14", 
         "Általános adatok", "Allgemeine Daten", "5", "", "", "", "", ""],
        ["6", "Szintek száma", "Anzahl Haltestellen", "number", "TRUE", "N14", 
         "Általános adatok", "Allgemeine Daten", "6", "", "", "", "", ""],
        ["7", "Otis Lift-azonosító", "OTIS Lift-Identifikation", "text", "TRUE", "O16", 
         "Általános adatok", "Allgemeine Daten", "7", "", "", "", "", ""],
        ["8", "Gyári szám", "Fabriknummer", "text", "TRUE", "O17", 
         "Általános adatok", "Allgemeine Daten", "8", "", "", "", "", ""],
        ["9", "Megjegyzés", "Bemerkung", "text", "FALSE", "O19", 
         "Általános adatok", "Allgemeine Daten", "9", "", "", "", "", ""],
        
        # Yes/No/NA kérdések
        ["10", "Bejelentés megtörtént", "Anmeldung erfolgt", "yes_no_na", "TRUE", 
         "A75;A76;A77,B75;B76;B77,C75;C76;C77", "Adminisztratív", "Administrativ", "1", "", "", "", "", ""],
        ["11", "Műszaki átadás megtörtént", "Technische Übergabe erfolgt", "yes_no_na", "TRUE", 
         "A79;A80;A81,B79;B80;B81,C79;C80;C81", "Adminisztratív", "Administrativ", "2", "", "", "", "", ""],
        
        # True/False kérdések  
        ["12", "Invertercsere", "Frequenzumrichter-Austausch", "true_false", "TRUE", "Q25", 
         "Modernizációban érintett", "Modernisierung betroffen", "1", "", "", "", "", ""],
        ["13", "Kabincsere", "Kabinen-Austausch", "true_false", "TRUE", "Q26", 
         "Modernizációban érintett", "Modernisierung betroffen", "2", "", "", "", "", ""],
        ["14", "Ajtómotor csere", "Türantrieb-Austausch", "true_false", "TRUE", "Q27", 
         "Modernizációban érintett", "Modernisierung betroffen", "3", "", "", "", "", ""],
        ["15", "Új vezérlőszekrény", "Neuer Steuerschrank", "true_false", "TRUE", "Q28", 
         "Modernizációban érintett", "Modernisierung betroffen", "4", "", "", "", "", ""],
        ["16", "Teljes új vezérlés tabló, külső hívók, stb.", "Komplett neue Steuerung, Tableaus, Ruftaster usw.", "true_false", "TRUE", "Q29", 
         "Modernizációban érintett", "Modernisierung betroffen", "5", "", "", "", "", ""],
        ["17", "Motor / Drótkötelek  /Szijj", "Motor / Drahtseile / Riemen", "true_false", "TRUE", "Q30", 
         "Modernizációban érintett", "Modernisierung betroffen", "6", "", "", "", "", ""],
        ["18", "Sebességszabályzó csere", "Geschwindigkeitsbegrenzer-Austausch", "true_false", "TRUE", "Q31", 
         "Modernizációban érintett", "Modernisierung betroffen", "7", "", "", "", "", ""],
        ["19", "Kabin és kabinajtó", "Kabine und Kabinentür", "true_false", "TRUE", "Q32", 
         "Modernizációban érintett", "Modernisierung betroffen", "8", "", "", "", "", ""],
        ["20", "Kézi-aknaajtó zárcsere", "Handbetätigung Schachttürschloss-Austausch", "true_false", "TRUE", "Q33", 
         "Modernizációban érintett", "Modernisierung betroffen", "9", "", "", "", "", ""],
        ["21", "Aknaajtók", "Schachttüren", "true_false", "TRUE", "Q34", 
         "Modernizációban érintett", "Modernisierung betroffen", "10", "", "", "", "", ""],
        
        # Mérési kérdések
        ["m1", "Távolság kabintető és Aknatető között", "Abstand zwischen Kabinendach und Schachtkopf", "measurement", "TRUE", "I278", 
         "Mérési adatok", "Messdaten", "1", "mm", "0", "10000", "", ""],
        ["m2", "Távolság kabintető legmagasabb pontja és Aknatető között", "Abstand zwischen höchstem Punkt Kabinendach und Schachtkopf", "measurement", "TRUE", "N278", 
         "Mérési adatok", "Messdaten", "2", "mm", "0", "10000", "", ""],
        ["m3", "Távolság az akna padló és az ellensúly puffer között", "Abstand zwischen Schachtgrube und Gegengewichtpuffer", "measurement", "TRUE", "I280", 
         "Mérési adatok", "Messdaten", "3", "mm", "0", "10000", "", ""],
        
        # Számított kérdések
        ["m4", "Effektív távolság A", "Effektive Entfernung A", "calculated", "TRUE", "I283", 
         "Számított értékek", "Berechnete Werte", "1", "mm", "700", "9000", "m1 - m3", "m1,m3"],
        ["m5", "Effektív távolság B", "Effektive Entfernung B", "calculated", "TRUE", "N283", 
         "Számított értékek", "Berechnete Werte", "2", "mm", "400", "9000", "m2 - m3", "m2,m3"],
    ]
    
    # Adatok hozzáadása
    for row_idx, question in enumerate(questions, 2):
        for col_idx, value in enumerate(question, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Oszlopszélesség beállítása
    column_widths = [8, 25, 25, 12, 10, 15, 20, 20, 8, 8, 10, 10, 20, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Protokoll template sheet hozzáadása (minta)
    protocol_sheet = wb.create_sheet("protocol")
    protocol_sheet["A1"] = "OTIS Átvételi Protokoll Template"
    protocol_sheet["A1"].font = Font(size=16, bold=True)
    
    # Minta cellák ahol az adatok megjelennek majd
    protocol_sheet["F9"] = "{{ Átvevő neve }}"
    protocol_sheet["Q9"] = "{{ Létesítmény címe }}"
    protocol_sheet["I278"] = "{{ m1 érték }}"
    protocol_sheet["N278"] = "{{ m2 érték }}"
    protocol_sheet["I280"] = "{{ m3 érték }}"
    protocol_sheet["I283"] = "{{ m4 számított érték }}"
    protocol_sheet["N283"] = "{{ m5 számított érték }}"
    
    # Fájl mentése
    wb.save("EGYESÍTETT-TEMPLATE-FULL.xlsx")
    print("✅ Egyesített template létrehozva: EGYESÍTETT-TEMPLATE-FULL.xlsx")
    print("")
    print("Ez a template tartalmazza:")
    print("📝 Szöveges és numerikus kérdéseket")
    print("✅ Yes/No/NA választó kérdéseket")  
    print("🔘 True/False kapcsoló kérdéseket")
    print("📏 Mérési kérdéseket (unit, min, max értékekkel)")
    print("🧮 Számított kérdéseket (formulák és input hivatkozásokkal)")
    print("")
    print("Minden kérdéstípus egy Excel fájlban van!")

if __name__ == "__main__":
    create_unified_template()